from openpyxl import *
from tkinter import *
import sys
import os

def restart_program():
    os.system('python "D:\PROJECTS\Apartment Management System\Test\Test_app.py"')

wb = Workbook()
ws = wb.active
book = load_workbook("Bill_Database.xlsx")
sheet = book['Sheet']

class Bill_database_create:

    def add_title(self):
        sheet['A1'] = "Serial"
        sheet['B1'] = "Flat No."
        sheet['C1'] = "Owner's Name"
        sheet['D1'] = "Electricity Bill"
        sheet['E1'] = "Gas Bill"
        sheet['F1'] = "Water Bill"
        sheet['G1'] = "Present Monthly bill"

    def save_excel(self):
        book.save("Bill_Database.xlsx")


class Bill_database_edit:

    def delete_sheet(std):
        sheet.remove_sheet(sheet.get_sheet_by_name(std))

    def add_information(self, member_list):
        sheet.append([sheet.max_row,member_list[0], member_list[1], member_list[2], member_list[3], member_list[4]])

    def delete_information_row(self, row):
        sheet.delete_rows(row+1)

    def delete_information_column(self, column):
        sheet.delete_cols(column)

    def delete_row_in_range(self, row1, row2):
        sheet.delete_rows(row1+1, row2-1)

    def delete_column_in_range(self, column1, column2):
        sheet.delete_cols(column1, column2)




class Bill_database_read:

    def print_database(self):
        for i in range(sheet.max_row):
            for j in range(sheet.max_column):
                print(sheet.cell(i+1,j+1).value,"  ", end="")
            print()

    def load_workbook_values(self):
        Bill_data_list = []
        for i in range(sheet.max_row):
            singleArray = []
            for j in range(sheet.max_column):
                singleArray.append(sheet.cell(i + 1, j + 1).value)
            Bill_data_list.append(singleArray)
        return Bill_data_list




class Bill_Calculation:

    def monthly_calculation(self):
        for row in range(2, sheet.max_row+1):
            sum = 0
            for column in range(4,7):
                sum += int(sheet.cell(row,column).value)
            sheet.cell(row,7).value=sum

    def electric_Bill(self):
        sum=0
        for row in range(2, sheet.max_row+1):
            sum+=int(sheet.cell(row,4).value)
        # sheet.cell(row+1, 3).value = sum
        return sum


    def gas_Bill(self):
        sum=0
        for row in range(2, sheet.max_row+1):
            sum+=int(sheet.cell(row,5).value)
        # sheet.cell(row+1, 4).value = sum
        return sum

    def water_bill(self):
        sum=0
        for row in range(2, sheet.max_row+1):
            sum+=int(sheet.cell(row,6).value)
        # sheet.cell(row+1, 5).value = sum
        return sum

    def total_monthly_bill(self):
        sum=0
        for row in range(2, sheet.max_row+1):
            if sheet.cell(row,7).value != None:
                sum+=int(sheet.cell(row,7).value)
            else:
                break
        # sheet.cell(row+1, 5).value = sum
        return sum



book2 = load_workbook("Flat_Database.xlsx")
sheet2 = book2['Sheet']

class Flat_database_create:
    # def __init__(self):
    #     self.wb = load_workbook("Flat_Database.xlsx")
    #     self.sheet = self.wb['Sheet']

    def add_title(self):
        sheet2['A1'] = "Serial"
        sheet2['B1'] = "Flat No."
        sheet2['C1'] = "Owner's Name"
        sheet2['D1'] = "To-Let"
        sheet2['E1'] = "Parking"
        sheet2['F1'] = "Bill Due"
        sheet2['G1'] = "Service Charge"
        sheet2['H1'] = "Contact Number"
        sheet2['I1'] = "Intercom Number"

    def save_excel(self):
        book2.save("Flat_Database.xlsx")


class Flat_database_edit:

    def add_information(self, member_list):
        sheet2.append([sheet2.max_row, member_list[0], member_list[1], "No", member_list[4], "No", "Cleared", member_list[2], member_list[3]])

    def delete_information_row(self, row):
        sheet2.delete_rows(row+1)

    def delete_information_column(self, column):
        sheet2.delete_cols(column)

    def delete_row_in_range(self, row1, row2):
        sheet2.delete_rows(row1+1, row2+1)

    def delete_column_in_range(self, column1, column2):
        sheet2.delete_cols(column1, column2)


class Flat_database_read:

    def print_database(self):
        for i in range(sheet.max_row):
            for j in range(sheet.max_column):
                print(sheet.cell(i+1,j+1).value, "  ", end="")
            print()

    def load_workbook_values(self):
        flat_data_list = []
        for i in range(sheet2.max_row):
            singleArray=[]
            for j in range(sheet2.max_column):
                singleArray.append(sheet2.cell(i+1,j+1).value)
            flat_data_list.append(singleArray)
        return flat_data_list


#------------------------------------------------Graphical User Interphase-----------------------------------------------


class FlatProfile:
    def __init__(self, root):
        self.root=root
        Label(root, text="        ", fg='black').grid(row=0, column=0)
        Label(root, text="        ", fg='black').grid(row=1, column=0)
        Label(root, text="        ", fg='black').grid(row=2, column=0)
        Label(root, text="        ", fg='black').grid(row=3, column=0)

        # ------------------Title----------------------
        Label(root, text="FLAT PROFILE", fg='Black', font=("Helvetica", 15,'bold')).place(x=200, y=15)
        Label(root, text="Click flat number for details of specific flat information:", fg='Black', font=("Helvetica", 10)).place(x=0, y=60)

        refresh_button = Button(root, text="Refresh", command=self.refresh)
        refresh_button.place(x=10, y=20)

        delete_flat_row_button = Button(root, text="Delete row :", command=self.delete_bill_row)
        delete_flat_row_button.place(x=610, y=20)
        self.delete_flat_row_value = Entry(width=5)
        self.delete_flat_row_value.place(x=690, y=23)


        Home_button = Button(root, text="Home Page", command=self.goto_home_page)
        Home_button.place(x=400, y=20)

        quit_button = Button(root, text="Quit", command=root.destroy)
        quit_button.place(x=490, y=20)

        for i in range(self.total_rows):
            for j in range(self.total_columns):
                if i==0:
                    self.e = Label(root, text=str(self.flat_data_list[i][j])+"    ", fg='black', font=("Helvetica", 10,'bold'))
                    self.e.grid(row=i+4, column=j)
                else:
                    if j==1:
                        self.btn = Button(root, text=self.flat_data_list[i][j],  width=5)
                        self.btn.grid(row=i+5, column=j, pady=10)
                    else:
                        self.e = Label(root, text=self.flat_data_list[i][j], fg='black').grid(row=i+5, column=j)

    def goto_home_page(self):
        self.root.destroy()
        present_Homepage_gui_frame()

    def refresh(self):
        self.root.destroy()
        restart_program()


    def delete_bill_row(self):
        Flat_database_edit().delete_information_row(int(self.delete_flat_row_value.get()))
        Flat_database_create().save_excel()

        self.root.destroy()
        restart_program()

    flat_data_list = Flat_database_read().load_workbook_values()

    total_rows = len(flat_data_list)
    total_columns = len(flat_data_list[0])





class Bill_Section:
    def __init__(self, root):

        self.root = root

        Label(root, text="        ", fg='black').grid(row=0, column=0)
        Label(root, text="        ", fg='black').grid(row=1, column=0)
        Label(root, text="        ", fg='black').grid(row=2, column=0)

        # ------------------Title----------------------
        Label(root, text="BILL SECTION", fg='Black', font=("Helvetica", 15, 'bold')).place(x=60, y=10)
        Label(root, text="Click flat number & bill titles for detailed information & money voucher:", fg='Black', font=("Helvetica", 10,)).place(x=0, y=50)

        refresh_button = Button(root, text="Refresh", command=self.refresh)
        refresh_button.place(x=700, y=140)

        Home_button = Button(root, text="Home Page", command=self.goto_home_page)
        Home_button.place(x=295, y=20)

        quit_button = Button(root, text="Quit", command=root.destroy)
        quit_button.place(x=260, y=20)

        generate_monthly_bill = Button(root, text="Generate monthly bill each", command=self.monthly_values)
        generate_monthly_bill.place(x=700, y=100)

        delete_bill_row_button = Button(root, text="Delete row :", command=self.delete_bill_row)
        delete_bill_row_button.place(x=610, y=20)
        self.delete_bill_row_value = Entry(width=5)
        self.delete_bill_row_value.place(x=690, y=23)

        delete_bill_row_button_range = Button(root, text="Delete row in range :", command=self.delete_bill_row_in_range)
        delete_bill_row_button_range.place(x=380, y=20)

        self.delete_bill_row_value_from = Entry(width=3)
        self.delete_bill_row_value_from.place(x=505, y=23)

        Label(root, text="to", fg ="Black").place(x=527, y=23)

        self.delete_bill_row_value_to = Entry(width=3)
        self.delete_bill_row_value_to.place(x=545, y=23)



        for i in range(self.total_rows):
            for j in range(self.total_columns):
                if i == 0:
                    if j==3 or j==4 or j==5:
                        self.btn = Button(root, text=str(self.bill_data_list[i][j])+"   ", font=("Helvetica", 9, 'bold'), pady=4)
                        self.btn.grid(row=i + 3, column=j, pady=10)
                    else:
                        self.e = Label(root, text=str(self.bill_data_list[i][j])+"   ", fg='black', font=("Helvetica", 10, 'bold'))
                        self.e.grid(row=i + 3, column=j)
                else:
                    if j == 1:
                        self.btn = Button(root, text=self.bill_data_list[i ][j], width=5)
                        self.btn.grid(row=i + 3, column=j, pady=10)
                    else:
                        self.e = Label(root, text=self.bill_data_list[i][j], fg='black').grid(row=i + 3, column=j, pady=10)

                if i==self.total_rows-1:
                    if j==2:
                        Label(root, text='TOTAL = ', fg='black', font=("Helvetica", 9, 'bold'), pady=20).grid(row=i+5, column=j)
                    elif j==3:
                        Label(root, text=str(Bill_Calculation().electric_Bill()), fg='black', font=("Helvetica", 9, 'bold'), pady=20).grid(row=i + 5, column=j)
                    elif j==4:
                        Label(root, text=str(Bill_Calculation().gas_Bill()), fg='black', font=("Helvetica", 9, 'bold'), pady=20).grid(row=i + 5, column=j)
                    elif j==5:
                        Label(root, text=str(Bill_Calculation().water_bill()), fg='black', font=("Helvetica", 9, 'bold'), pady=20).grid(row=i + 5, column=j)
                    elif j==6:
                        Label(root, text=str(Bill_Calculation().total_monthly_bill()), fg='black', font=("Helvetica", 9, 'bold'), pady=20).grid(row=i + 5, column=j)

    def monthly_values(self):
        Bill_Calculation().monthly_calculation()
        Bill_database_create().save_excel()

    def goto_home_page(self):
        self.root.destroy()
        present_Homepage_gui_frame()

    def delete_bill_row(self):
        Bill_database_edit().delete_information_row(int(self.delete_bill_row_value.get()))
        Bill_database_create().save_excel()

        self.root.destroy()
        restart_program()

    def delete_bill_row_in_range(self):
        row_from = int(self.delete_bill_row_value_from.get())
        row_to = int(self.delete_bill_row_value_to.get())
        Bill_database_edit().delete_row_in_range(row_from, row_to)
        Bill_database_create().save_excel()

    def refresh(self):
        self.root.destroy()
        restart_program()

    bill_data_list = Bill_database_read().load_workbook_values()
    total_rows = len(bill_data_list)
    total_columns = len(bill_data_list[0])




class Add_Bill:
    def __init__(self, master):

        self.master = master

        Label(master, text="                             ", font=("Helvetica", 10)).grid(row=0, column=0)
        Label(master, text="Add Bill info-", font=("Helvetica", 10,'bold')).place(x=0, y=0)

        # refresh_button = Button(master, text="Refresh", command=self.refresh)
        # refresh_button.grid(x=10, column=3)

        Home_button = Button(master, text="Home Page", command=self.goto_home_page)
        Home_button.grid(row=10, column=0, sticky=E)

        quit_button = Button(master, text="Quit", command=master.destroy)
        quit_button.grid(row=10, column=2, sticky=E)

        l1 = Label(master, text="Flat number :")
        l2 = Label(master, text="Owner name :")
        l3 = Label(master, text="Electric Bill :")
        l4 = Label(master, text="Gas Bill:")
        l5 = Label(master, text="Water Bill:")

        l1.grid(row=1, column=0, sticky=W, pady=2)
        l2.grid(row=2, column=0, sticky=W, pady=2)
        l3.grid(row=3, column=0, sticky=W, pady=2)
        l4.grid(row=4, column=0, sticky=W, pady=2)
        l5.grid(row=5, column=0, sticky=W, pady=2)

        self.e1 = Entry()
        self.e2 = Entry()
        self.e3 = Entry()
        self.e4 = Entry()
        self.e5 = Entry()

        self.e1.grid(row=1, column=1, pady=2)
        self.e2.grid(row=2, column=1, pady=2)
        self.e3.grid(row=3, column=1, pady=2)
        self.e4.grid(row=4, column=1, pady=2)
        self.e5.grid(row=5, column=1, pady=2)

        b1 = Button(master, text="Add", command=self.add_value)
        b1.grid(row=10, column=1, sticky=E)


    def add_value(self):
        data_list = []
        data_list.append(self.e1.get())
        data_list.append(self.e2.get())
        data_list.append(self.e3.get())
        data_list.append(self.e4.get())
        data_list.append(self.e5.get())
        Bill_database_edit().add_information(data_list)
        Bill_database_create().save_excel()
        # print(self.e1.get(), self.e2.get(), self.e3.get(), self.e4.get(), self.e5.get())

        self.master.destroy()
        restart_program()

    def goto_home_page(self):
        self.master.destroy()
        present_Homepage_gui_frame()




class Add_Owner:
    def __init__(self, master):

        self.master = master

        Label(master, text="                             ", font=("Helvetica", 10)).grid(row=0, column=0)
        Label(master, text="Add flat owner info-", font=("Helvetica", 10,'bold')).place(x=0, y=0)

        Home_button = Button(master, text="Home Page", command=self.goto_home_page)
        Home_button.grid(row=10, column=0, sticky=E)

        quit_button = Button(master, text="Quit", command=master.destroy)
        quit_button.grid(row=10, column=2, sticky=E)

        # this will create a label widget
        l1 = Label(master, text="Flat number :")
        l2 = Label(master, text="Owner name :")
        l3 = Label(master, text="Contact Number :")
        l4 = Label(master, text="Intercom no. :")
        l5 = Label(master, text="Parking :")

        l1.grid(row=1, column=0, sticky=W, pady=2)
        l2.grid(row=2, column=0, sticky=W, pady=2)
        l3.grid(row=3, column=0, sticky=W, pady=2)
        l4.grid(row=4, column=0, sticky=W, pady=2)
        l5.grid(row=5, column=0, sticky=W, pady=2)

        # entry widgets, used to take entry from user
        self.e1 = Entry()
        self.e2 = Entry()
        self.e3 = Entry()
        self.e4 = Entry()
        self.e5 = Entry()

        # this will arrange entry widgets
        self.e1.grid(row=1, column=1, pady=2)
        self.e2.grid(row=2, column=1, pady=2)
        self.e3.grid(row=3, column=1, pady=2)
        self.e4.grid(row=4, column=1, pady=2)
        self.e5.grid(row=5, column=1, pady=2)

        # # button widget
        b1 = Button(master, text="Add", command=self.add_value)
        b1.grid(row=10, column=1, sticky=E)


    def add_value(self):
        data_list=[]
        data_list.append(self.e1.get())
        data_list.append(self.e2.get())
        data_list.append(self.e3.get())
        data_list.append(self.e4.get())
        data_list.append(self.e5.get())
        Flat_database_edit().add_information(data_list)
        Flat_database_create().save_excel()
        # print(self.e1.get(), self.e2.get(), self.e3.get(), self.e4.get(), self.e5.get())

        self.master.destroy()
        restart_program()

    def goto_home_page(self):
        self.master.destroy()
        present_Homepage_gui_frame()




class Homepage:
    def __init__(self, window):
        self.window = window
        #------------------Title----------------------
        Title=Label(window, text="HOME PAGE", fg='Black', font=("Helvetica", 40,'bold'))
        Title.place(x=220, y=10)

        note = Label(window, text="Add 2 excel files named 'Flat_Database' & 'Bill_database' in the folder of the app to run and click the above right buttons to create titles in the database", fg='Black', font=("Helvetica", 9))
        note.place(x=10, y=400)

        create_Flat_data_button = Button(window, text="Create title for flat profile", command=self.createFlatDatabase)
        create_Flat_data_button.place(x=400, y=200)

        create_bill_data_button = Button(window, text="Create title for bill section", command=self.createBilldatabase)
        create_bill_data_button.place(x=400, y=240)

        quit_button = Button(window, text="Quit", command=window.destroy)
        quit_button.place(x=400, y=300)

        #-----------------Sections---------------------
        lbl=Label(window, text="Select a database to visit in the list below:", fg='Black', font=("Helvetica", 14,'bold'))
        lbl.place(x=10, y=100)

        flat_Profile_btn = Button(window, text="Flat Profile", fg='blue', command=self.goto_flatProfile)
        flat_Profile_btn.place(x=20, y=140)

        Bill_Section_btn = Button(window, text="Bill Section", fg='blue', command=self.goto_billSection)
        Bill_Section_btn.place(x=20, y=180)

        Add_bill_btn = Button(window, text="Add Bill", fg='blue', command=self.goto_AddBill)
        Add_bill_btn.place(x=20, y=220)

        Add_flat_members_btn = Button(window, text="Add flat members", fg='blue', command=self.goto_AddMember)
        Add_flat_members_btn.place(x=20, y=260)

        Service_Charge_btn = Button(window, text="Service Charge", fg='blue')
        Service_Charge_btn.place(x=20, y=300)

        Maintenance_btn = Button(window, text="Maintenance", fg='blue')
        Maintenance_btn.place(x=20, y=340)

    def createFlatDatabase(self):
        Flat_database_create().add_title()
        Flat_database_create().save_excel()

    def createBilldatabase(self):
        Bill_database_create().add_title()
        Bill_database_create().save_excel()

    def goto_flatProfile(self):
        self.window.destroy()
        Controller().open_flatProfile_gui()

    def goto_billSection(self):
        self.window.destroy()
        Controller().open_bill_section_gui()

    def goto_AddMember(self):
        self.window.destroy()
        Controller().open_addMember_gui()

    def goto_AddBill(self):
        self.window.destroy()
        Controller().open_addBill_gui()



class Controller:
    # Bill_database_read().print_database()
    # Flat_database_read().print_database()
    def open_Application(self):
        present_Homepage_gui_frame()

    def open_bill_section_gui(self):
        present_BillSection_gui_frame()

    def open_addMember_gui(self):
        present_AddOwner_gui_frame()

    def open_addBill_gui(self):
        present_AddBill_gui_frame()

    def open_flatProfile_gui(self):
        present_FlatProfile_gui_frame()



def present_Homepage_gui_frame():
    window = Tk()
    Homepage(window)
    window.title('Apartment Management System')
    window.geometry("840x450+10+20")
    window.mainloop()


def present_AddOwner_gui_frame():
    master = Tk()
    Add_Owner(master)
    master.title("Apartment Management System")
    master.geometry("400x200+10+20")
    mainloop()


def present_AddBill_gui_frame():
    master = Tk()
    Add_Bill(master)
    master.title("Apartment Management System")
    master.geometry("320x200+10+10")
    mainloop()


def present_BillSection_gui_frame():
    root = Tk()
    Bill_Section(root)
    root.title('Apartment Management System')
    root.geometry("950x500+10+20")
    root.mainloop()


def present_FlatProfile_gui_frame():
    root = Tk()
    FlatProfile(root)
    root.title('Apartment Management System')
    root.geometry("860x350+10+20")
    root.mainloop()


Controller().open_Application()