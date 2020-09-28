from openpyxl import *
import os

wb = Workbook()
ws = wb.active

try:
    book = load_workbook("Bill_Database.xlsx")
    sheet = book['Sheet']

except:
    file01 = os.getcwd() + "\Bill_Database.xlsx"

    wb.save(file01)

    book = load_workbook("Bill_Database.xlsx")
    sheet = book['Sheet']
    print("New Files Created")


class Bill_database_create:

    @staticmethod
    def add_title():
        sheet['A1'] = "Serial"
        sheet['B1'] = "Flat No."
        sheet['C1'] = "Owner's Name"
        sheet['D1'] = "Electricity Bill"
        sheet['E1'] = "Gas Bill"
        sheet['F1'] = "Water Bill"
        sheet['G1'] = "Present Monthly bill"

    @staticmethod
    def save_excel():
        book.save("Bill_Database.xlsx")


class Bill_database_edit:

    @staticmethod
    def add_information(member_list):
        if len(member_list) > 2:
            r = ""
            for i in range(1, sheet.max_row):
                if member_list[0] == sheet.cell(i + 1, 2).value:
                    r = i + 1
                    print("match")
                    break

            sheet.cell(r, 4).value = member_list[2]
            sheet.cell(r, 5).value = member_list[3]
            sheet.cell(r, 6).value = member_list[4]

        else:
            sheet.append([sheet.max_row, member_list[0], member_list[1], 0, 0, 0])

    @staticmethod
    def delete_information_row(row):
        sheet.delete_rows(row + 1)

    @staticmethod
    def make_it_zero_row(row):
        sheet.cell(row + 1, 4).value = 0
        sheet.cell(row + 1, 5).value = 0
        sheet.cell(row + 1, 6).value = 0
        sheet.cell(row + 1, 7).value = 0

    @staticmethod
    def delete_information_column(column):
        sheet.delete_cols(column)

    @staticmethod
    def delete_row_in_range(row1, row2):
        sheet.delete_rows(row1 + 1, row2 - 1)

    @staticmethod
    def delete_column_in_range(column1, column2):
        sheet.delete_cols(column1, column2)

    @staticmethod
    def delete_gas_bill_row(row):
        sheet.cell(row + 1, 5).value = 0

    @staticmethod
    def delete_electric_bill_row(row):
        sheet.cell(row + 1, 4).value = 0

    @staticmethod
    def delete_water_bill(row):
        sheet.cell(row + 1, 6).value = 0


class Bill_database_read:

    @staticmethod
    def print_database():
        for i in range(sheet.max_row):
            for j in range(sheet.max_column):
                print(sheet.cell(i + 1, j + 1).value, "  ", end="")
            print()

    @staticmethod
    def load_workbook_values():
        Bill_data_list = []
        for i in range(sheet.max_row):
            singleArray = []
            for j in range(sheet.max_column):
                singleArray.append(sheet.cell(i + 1, j + 1).value)
            Bill_data_list.append(singleArray)
        return Bill_data_list


class Bill_Calculation:

    @staticmethod
    def monthly_calculation():
        for row in range(2, sheet.max_row + 1):
            sum = 0
            for column in range(4, 7):
                sum += int(sheet.cell(row, column).value)
            sheet.cell(row, 7).value = sum

    @staticmethod
    def electric_Bill():
        sum = 0
        for row in range(2, sheet.max_row + 1):
            sum += int(sheet.cell(row, 4).value)
        return sum

    @staticmethod
    def gas_Bill():
        sum = 0
        for row in range(2, sheet.max_row + 1):
            sum += int(sheet.cell(row, 5).value)
        return sum

    @staticmethod
    def water_bill():
        sum = 0
        for row in range(2, sheet.max_row + 1):
            sum += int(sheet.cell(row, 6).value)
        return sum

    @staticmethod
    def total_monthly_bill():
        sum = 0
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row, 7).value is not None:
                sum += int(sheet.cell(row, 7).value)
            else:
                break
        return sum
