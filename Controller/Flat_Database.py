from openpyxl import *
import os

wb = Workbook()
ws = wb.active

try:
    book2 = load_workbook("Flat_Database.xlsx")
    sheet2 = book2['Sheet']

except:
    file02 = os.getcwd() + "\Flat_Database.xlsx"

    wb.save(file02)
    book2 = load_workbook("Flat_Database.xlsx")
    sheet2 = book2['Sheet']
    print("New Files Created")


class Flat_database_create:

    @staticmethod
    def add_title():
        sheet2['A1'] = "Serial"
        sheet2['B1'] = "Flat No."
        sheet2['C1'] = "Owner's Name"
        sheet2['D1'] = "To-Let"
        sheet2['E1'] = "Parking"
        sheet2['F1'] = "Bill Due"
        sheet2['G1'] = "Service Charge"
        sheet2['H1'] = "Contact Number"
        sheet2['I1'] = "Intercom Number"

    @staticmethod
    def save_excel():
        book2.save("Flat_Database.xlsx")


class Flat_database_edit:

    @staticmethod
    def add_information(member_list):
        sheet2.append(
            [sheet2.max_row, member_list[0], member_list[1], "No", member_list[4], "No", "Cleared", member_list[2],
             member_list[3]])

    @staticmethod
    def delete_information_row(row):
        sheet2.delete_rows(row + 1)

    @staticmethod
    def delete_information_column(column):
        sheet2.delete_cols(column)

    @staticmethod
    def delete_row_in_range(row1, row2):
        sheet2.delete_rows(row1 + 1, row2 + 1)

    @staticmethod
    def delete_column_in_range(column1, column2):
        sheet2.delete_cols(column1, column2)


class Flat_database_read:

    @staticmethod
    def print_database():
        for i in range(sheet2.max_row):
            for j in range(sheet2.max_column):
                print(sheet2.cell(i + 1, j + 1).value, "  ", end="")
            print()

    @staticmethod
    def load_workbook_values():
        flat_data_list = []
        for i in range(sheet2.max_row):
            singleArray = []
            for j in range(sheet2.max_column):
                singleArray.append(sheet2.cell(i + 1, j + 1).value)
            flat_data_list.append(singleArray)
        return flat_data_list

# workbook_read = Flat_database_read()
# workbook_read.load_workbook_values()
# workbook_read.save_excel()
