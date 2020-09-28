from openpyxl import *
from tkinter import *
import os


def restart_program():
    os.execv(sys.executable, [sys.executable, '"' + sys.argv[0] + '"'] + sys.argv[1:])


wb = Workbook()
ws = wb.active

try:
    book = load_workbook("Bill_Database.xlsx")
    sheet = book['Sheet']
    book2 = load_workbook("Flat_Database.xlsx")
    sheet2 = book2['Sheet']

except:
    file01 = os.getcwd() + "\Bill_Database.xlsx"
    file02 = os.getcwd() + "\Flat_Database.xlsx"

    wb.save(file01)
    wb.save(file02)

    book = load_workbook("Bill_Database.xlsx")
    sheet = book['Sheet']
    book2 = load_workbook("Flat_Database.xlsx")
    sheet2 = book2['Sheet']
    print("New Files Created")


class Bill_database_create:

    @staticmethod
    def add_title():
        sheet['A1'] = "Serial"
        sheet['B1'] = "Flat No."
        sheet['C1'] = "Owner's Name"
        sheet['D1'] = "Electricity Bill"
        sheet['E1'] = "Gas Bill"
        sheet['F1'] = "Water Bill"
        sheet['G1'] = "Present Monthly bill"

    @staticmethod
    def save_excel():
        book.save("Bill_Database.xlsx")


class Bill_database_edit:

    @staticmethod
    def add_information(member_list):
        if len(member_list) > 2:
            r = ""
            for i in range(1, sheet.max_row):
                if member_list[0] == sheet.cell(i + 1, 2).value:
                    r = i + 1
                    print("match")
                    break

            sheet.cell(r, 4).value = member_list[2]
            sheet.cell(r, 5).value = member_list[3]
            sheet.cell(r, 6).value = member_list[4]

        else:
            sheet.append([sheet.max_row, member_list[0], member_list[1], 0, 0, 0])

    @staticmethod
    def delete_information_row(row):
        sheet.delete_rows(row + 1)

    @staticmethod
    def make_it_zero_row(row):
        sheet.cell(row + 1, 4).value = 0
        sheet.cell(row + 1, 5).value = 0
        sheet.cell(row + 1, 6).value = 0
        sheet.cell(row + 1, 7).value = 0

    @staticmethod
    def delete_information_column(column):
        sheet.delete_cols(column)

    @staticmethod
    def delete_row_in_range(row1, row2):
        sheet.delete_rows(row1 + 1, row2 - 1)

    @staticmethod
    def delete_column_in_range(column1, column2):
        sheet.delete_cols(column1, column2)

    @staticmethod
    def delete_gas_bill_row(row):
        sheet.cell(row + 1, 5).value = 0

    @staticmethod
    def delete_electric_bill_row(row):
        sheet.cell(row + 1, 4).value = 0

    @staticmethod
    def delete_water_bill(row):
        sheet.cell(row + 1, 6).value = 0


class Bill_database_read:

    @staticmethod
    def print_database():
        for i in range(sheet.max_row):
            for j in range(sheet.max_column):
                print(sheet.cell(i + 1, j + 1).value, "  ", end="")
            print()

    @staticmethod
    def load_workbook_values():
        Bill_data_list = []
        for i in range(sheet.max_row):
            singleArray = []
            for j in range(sheet.max_column):
                singleArray.append(sheet.cell(i + 1, j + 1).value)
            Bill_data_list.append(singleArray)
        return Bill_data_list


class Bill_Calculation:

    @staticmethod
    def monthly_calculation():
        for row in range(2, sheet.max_row + 1):
            sum = 0
            for column in range(4, 7):
                sum += int(sheet.cell(row, column).value)
            sheet.cell(row, 7).value = sum

    @staticmethod
    def electric_Bill():
        sum = 0
        for row in range(2, sheet.max_row + 1):
            sum += int(sheet.cell(row, 4).value)
        return sum

    @staticmethod
    def gas_Bill():
        sum = 0
        for row in range(2, sheet.max_row + 1):
            sum += int(sheet.cell(row, 5).value)
        return sum

    @staticmethod
    def water_bill():
        sum = 0
        for row in range(2, sheet.max_row + 1):
            sum += int(sheet.cell(row, 6).value)
        return sum

    @staticmethod
    def total_monthly_bill():
        sum = 0
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row, 7).value is not None:
                sum += int(sheet.cell(row, 7).value)
            else:
                break
        return sum


class Flat_database_create:

    @staticmethod
    def add_title():
        sheet2['A1'] = "Serial"
        sheet2['B1'] = "Flat No."
        sheet2['C1'] = "Owner's Name"
        sheet2['D1'] = "To-Let"
        sheet2['E1'] = "Parking"
        sheet2['F1'] = "Bill Due"
        sheet2['G1'] = "Service Charge"
        sheet2['H1'] = "Contact Number"
        sheet2['I1'] = "Intercom Number"

    @staticmethod
    def save_excel():
        book2.save("Flat_Database.xlsx")


class Flat_database_edit:

    @staticmethod
    def add_information(member_list):
        sheet2.append(
            [sheet2.max_row, member_list[0], member_list[1], "No", member_list[4], "No", "Cleared", member_list[2],
             member_list[3]])

    @staticmethod
    def delete_information_row(row):
        sheet2.delete_rows(row + 1)

    @staticmethod
    def delete_information_column(column):
        sheet2.delete_cols(column)

    @staticmethod
    def delete_row_in_range(row1, row2):
        sheet2.delete_rows(row1 + 1, row2 + 1)

    @staticmethod
    def delete_column_in_range(column1, column2):
        sheet2.delete_cols(column1, column2)


class Flat_database_read:

    @staticmethod
    def print_database():
        for i in range(sheet.max_row):
            for j in range(sheet.max_column):
                print(sheet.cell(i + 1, j + 1).value, "  ", end="")
            print()

    @staticmethod
    def load_workbook_values():
        flat_data_list = []
        for i in range(sheet2.max_row):
            singleArray = []
            for j in range(sheet2.max_column):
                singleArray.append(sheet2.cell(i + 1, j + 1).value)
            flat_data_list.append(singleArray)
        return flat_data_list


# ------------------------------------------------Graphical User Interface-----------------------------------------------


class FlatProfile:
    def __init__(self, root):
        self.root = root
        Label(root, text="        ", bg="#CAE9F5", fg='black').grid(row=0, column=0)
        Label(root, text="        ", bg="#CAE9F5", fg='black').grid(row=1, column=0)
        Label(root, text="        ", bg="#CAE9F5", fg='black').grid(row=2, column=0)
        Label(root, text="        ", bg="#CAE9F5", fg='black').grid(row=3, column=0)

        # ------------------Title----------------------
        Label(root, text="FLAT PROFILE", bg="#CAE9F5", fg='Black', font=("Helvetica", 15, 'bold')).place(x=200, y=15)
        Label(root, text="Click flat number for details of specific flat information:", bg="#CAE9F5", fg='Black',
              font=("Helvetica", 10)).place(x=0, y=60)

        refresh_button = Button(root, text="Refresh", bg="#ADECDF", command=self.refresh)
        refresh_button.place(x=10, y=20)

        delete_flat_row_button = Button(root, text="Delete row :", bg="#ADECDF", command=self.delete_bill_row)
        delete_flat_row_button.place(x=610, y=20)
        self.delete_flat_row_value = Entry(width=5)
        self.delete_flat_row_value.place(x=690, y=23)

        home_button = Button(root, text="Home Page", bg="#ADECDF", command=self.goto_home_page)
        home_button.place(x=400, y=20)

        quit_button = Button(root, text="Quit", bg="#ADECDF", command=root.destroy)
        quit_button.place(x=490, y=20)

        for i in range(self.total_rows):
            for j in range(self.total_columns):
                if i == 0:
                    self.e = Label(root, text=str(self.flat_data_list[i][j]) + "    ", bg="#CAE9F5", fg='black',
                                   font=("Helvetica", 10, 'bold'))
                    self.e.grid(row=i + 4, column=j)
                else:
                    if j == 1:
                        self.btn = Button(root, text=self.flat_data_list[i][j], width=5)
                        self.btn.grid(row=i + 5, column=j, pady=10)
                    else:
                        self.e = Label(root, text=self.flat_data_list[i][j], fg='black').grid(row=i + 5, column=j)

    def goto_home_page(self):
        self.root.destroy()
        present_Homepage_gui_frame()

    @staticmethod
    def refresh():
        restart_program()

    def delete_bill_row(self):
        Flat_database_edit().delete_information_row(int(self.delete_flat_row_value.get()))
        Flat_database_create().save_excel()

        Bill_database_edit().delete_information_row(int(self.delete_flat_row_value.get()))
        Bill_database_create().save_excel()

        self.refresh()

    flat_data_list = Flat_database_read().load_workbook_values()
    total_rows = len(flat_data_list)
    total_columns = len(flat_data_list[0])


class Bill_Section:
    def __init__(self, root):

        self.root = root

        Label(root, text="        ", bg="#CAE9F5", fg='black').grid(row=0, column=0)
        Label(root, text="        ", bg="#CAE9F5", fg='black').grid(row=1, column=0)
        Label(root, text="        ", bg="#CAE9F5", fg='black').grid(row=2, column=0)

        # ------------------Title----------------------
        Label(root, text="BILL SECTION", bg="#CAE9F5", fg='Black', font=("Helvetica", 15, 'bold')).place(x=60, y=10)
        Label(root, text="Click flat number & bill titles for detailed information & money voucher:", bg="#CAE9F5",
              fg='Black',
              font=("Helvetica", 10,)).place(x=0, y=50)

        refresh_button = Button(root, text="Refresh", bg="#ADECDF", command=self.refresh)
        refresh_button.place(x=700, y=140)

        Home_button = Button(root, text="Home Page", bg="#ADECDF", command=self.goto_home_page)
        Home_button.place(x=295, y=20)

        quit_button = Button(root, text="Quit", bg="#ADECDF", command=root.destroy)
        quit_button.place(x=260, y=20)

        generate_monthly_bill = Button(root, text="Generate monthly bill each", bg="#ADECDF",
                                       command=self.monthly_values)
        generate_monthly_bill.place(x=700, y=100)

        delete_bill_row_button = Button(root, text="Clear values of row :", bg="#ADECDF", command=self.delete_bill_row)
        delete_bill_row_button.place(x=565, y=20)
        self.delete_bill_row_value = Entry(width=5)
        self.delete_bill_row_value.place(x=690, y=23)

        for i in range(self.total_rows):
            for j in range(self.total_columns):
                if i == 0:
                    if j == 3:
                        self.btn = Button(root, text=str(self.bill_data_list[i][j]) + "   ", bg="#ADECDF",
                                          font=("Helvetica", 9, 'bold'), pady=4, command=self.goto_electric_bill)
                        self.btn.grid(row=i + 3, column=j, pady=10)
                    elif j == 4:
                        self.btn = Button(root, text=str(self.bill_data_list[i][j]) + "   ", bg="#ADECDF",
                                          font=("Helvetica", 9, 'bold'), pady=4, command=self.goto_gas_bill)
                        self.btn.grid(row=i + 3, column=j, pady=10)
                    elif j == 5:
                        self.btn = Button(root, text=str(self.bill_data_list[i][j]) + "   ", bg="#ADECDF",
                                          font=("Helvetica", 9, 'bold'), pady=4, command=self.goto_water_bill)
                        self.btn.grid(row=i + 3, column=j, pady=10)
                    else:
                        self.e = Label(root, text=str(self.bill_data_list[i][j]) + "   ", bg="#CAE9F5", fg='black',
                                       font=("Helvetica", 10, 'bold'))
                        self.e.grid(row=i + 3, column=j)
                else:
                    if j == 1:
                        self.btn = Button(root, text=self.bill_data_list[i][j], width=5)
                        self.btn.grid(row=i + 3, column=j, pady=10)
                    else:
                        self.e = Label(root, text=self.bill_data_list[i][j], fg='black').grid(row=i + 3, column=j,
                                                                                              pady=10)

                if i == self.total_rows - 1:
                    if j == 2:
                        Label(root, text='TOTAL = ', fg='black', font=("Helvetica", 9, 'bold'), bg="#CAE9F5",
                              pady=20).grid(row=i + 5, column=j)
                    elif j == 3:
                        Label(root, text=str(Bill_Calculation().electric_Bill()), bg="#CAE9F5", fg='black',
                              font=("Helvetica", 9, 'bold'), pady=20).grid(row=i + 5, column=j)
                    elif j == 4:
                        Label(root, text=str(Bill_Calculation().gas_Bill()), bg="#CAE9F5", fg='black',
                              font=("Helvetica", 9, 'bold'),
                              pady=20).grid(row=i + 5, column=j)
                    elif j == 5:
                        Label(root, text=str(Bill_Calculation().water_bill()), bg="#CAE9F5", fg='black',
                              font=("Helvetica", 9, 'bold'), pady=20).grid(row=i + 5, column=j)
                    elif j == 6:
                        Label(root, text=str(Bill_Calculation().total_monthly_bill()), bg="#CAE9F5", fg='black',
                              font=("Helvetica", 9, 'bold'), pady=20).grid(row=i + 5, column=j)

    @staticmethod
    def monthly_values():
        Bill_Calculation().monthly_calculation()
        Bill_database_create().save_excel()

    def goto_home_page(self):
        self.root.destroy()
        present_Homepage_gui_frame()

    def goto_water_bill(self):
        self.root.destroy()
        present_waterBill_gui_frame()

    def goto_gas_bill(self):
        self.root.destroy()
        present_gasBill_gui_frame()

    def goto_electric_bill(self):
        self.root.destroy()
        present_electricBill_gui_frame()

    def delete_bill_row(self):
        row = int(self.delete_bill_row_value.get())
        Bill_database_edit().make_it_zero_row(row)
        Bill_database_create().save_excel()
        self.refresh()

    @staticmethod
    def refresh():
        restart_program()

    bill_data_list = Bill_database_read().load_workbook_values()
    total_rows = len(bill_data_list)
    total_columns = len(bill_data_list[0])


class Electric_Bill:
    def __init__(self, root):
        self.root = root
        Label(root, text="        ", bg="#CAE9F5", fg='black').grid(row=0, column=0)
        Label(root, text="        ", bg="#CAE9F5", fg='black').grid(row=1, column=0)
        Label(root, text="        ", bg="#CAE9F5", fg='black').grid(row=2, column=0)

        # ------------------Title----------------------
        Label(root, text="Electric Bill", bg="#CAE9F5", fg='Black', font=("Helvetica", 15, 'bold')).place(x=60, y=10)

        refresh_button = Button(root, text="Refresh", bg="#ADECDF", command=self.refresh)
        refresh_button.place(x=360, y=140)

        Back_button = Button(root, text="Back", bg="#ADECDF", command=self.go_back)
        Back_button.place(x=255, y=20)

        Home_button = Button(root, text="Home Page", bg="#ADECDF", command=self.goto_home_page)
        Home_button.place(x=295, y=20)

        quit_button = Button(root, text="Quit", bg="#ADECDF", command=root.destroy)
        quit_button.place(x=360, y=200)

        delete_bill_row_button = Button(root, text="Delete row :", bg="#ADECDF", command=self.delete_bill_row)
        delete_bill_row_button.place(x=400, y=20)
        self.delete_bill_row_value = Entry(width=5)
        self.delete_bill_row_value.place(x=480, y=23)

        for i in range(self.total_rows):
            for j in range(self.total_columns):
                if j == 0 or j == 1 or j == 3:
                    if i == 0:
                        self.e = Label(root, text=str(self.elecricbill_data_list[i][j]) + "   ", bg="#CAE9F5",
                                       fg='black',
                                       font=("Helvetica", 10, 'bold'))
                        self.e.grid(row=i + 3, column=j)
                    else:
                        if j == 1:
                            self.btn = Button(root, text=self.elecricbill_data_list[i][j], width=5)
                            self.btn.grid(row=i + 3, column=j, pady=10)
                        else:
                            self.e = Label(root, text=self.elecricbill_data_list[i][j], fg='black').grid(row=i + 3,
                                                                                                         column=j,
                                                                                                         pady=10)

                if i == self.total_rows - 1:
                    if j == 1:
                        Label(root, text='TOTAL = ', fg='black', font=("Helvetica", 9, 'bold'), bg="#CAE9F5",
                              pady=20).grid(row=i + 5, column=j)

                    elif j == 5:
                        Label(root, text=str(Bill_Calculation().electric_Bill()), bg="#CAE9F5", fg='black',
                              font=("Helvetica", 9, 'bold'), pady=20).grid(row=i + 5, column=3)

    def go_back(self):
        self.root.destroy()
        present_BillSection_gui_frame()

    def delete_bill_row(self):
        row = int(self.delete_bill_row_value.get())
        Bill_database_edit().delete_electric_bill_row(row)
        Bill_database_create().save_excel()
        self.refresh()

    def goto_home_page(self):
        self.root.destroy()
        present_Homepage_gui_frame()

    @staticmethod
    def refresh():
        restart_program()

    elecricbill_data_list = Bill_database_read().load_workbook_values()
    total_rows = len(elecricbill_data_list)
    total_columns = len(elecricbill_data_list[0])


class Water_bill:
    def __init__(self, root):
        self.root = root
        Label(root, text="        ", bg="#CAE9F5", fg='black').grid(row=0, column=0)
        Label(root, text="        ", bg="#CAE9F5", fg='black').grid(row=1, column=0)
        Label(root, text="        ", bg="#CAE9F5", fg='black').grid(row=2, column=0)

        # ------------------Title----------------------
        Label(root, text="Water Bill", bg="#CAE9F5", fg='Black', font=("Helvetica", 15, 'bold')).place(x=60, y=10)

        refresh_button = Button(root, text="Refresh", bg="#ADECDF", command=self.refresh)
        refresh_button.place(x=360, y=140)

        Back_button = Button(root, text="Back", bg="#ADECDF", command=self.go_back)
        Back_button.place(x=255, y=20)

        Home_button = Button(root, text="Home Page", bg="#ADECDF", command=self.goto_home_page)
        Home_button.place(x=295, y=20)

        quit_button = Button(root, text="Quit", bg="#ADECDF", command=root.destroy)
        quit_button.place(x=360, y=200)

        delete_bill_row_button = Button(root, text="Clear row :", bg="#ADECDF", command=self.delete_bill_row)
        delete_bill_row_button.place(x=400, y=20)
        self.delete_bill_row_value = Entry(width=5)
        self.delete_bill_row_value.place(x=480, y=23)

        for i in range(self.total_rows):
            for j in range(self.total_columns):
                if j == 0 or j == 1 or j == 5:
                    if i == 0:
                        self.e = Label(root, text=str(self.waterbill_data_list[i][j]) + "   ", bg="#CAE9F5", fg='black',
                                       font=("Helvetica", 10, 'bold'))
                        self.e.grid(row=i + 3, column=j)
                    else:
                        if j == 1:
                            self.btn = Button(root, text=self.waterbill_data_list[i][j], width=5)
                            self.btn.grid(row=i + 3, column=j, pady=10)
                        else:
                            self.e = Label(root, text=self.waterbill_data_list[i][j], fg='black').grid(row=i + 3,
                                                                                                       column=j,
                                                                                                       pady=10)

                if i == self.total_rows - 1:
                    if j == 1:
                        Label(root, text='TOTAL = ', fg='black', font=("Helvetica", 9, 'bold'), bg="#CAE9F5",
                              pady=20).grid(row=i + 5, column=j)
                    elif j == 3:
                        Label(root, text=str(Bill_Calculation().water_bill()), bg="#CAE9F5", fg='black',
                              font=("Helvetica", 9, 'bold'), pady=20).grid(row=i + 5, column=5)

    def go_back(self):
        self.root.destroy()
        present_BillSection_gui_frame()

    def delete_bill_row(self):
        row = int(self.delete_bill_row_value.get())
        Bill_database_edit().delete_water_bill(row)
        Bill_database_create().save_excel()
        self.refresh()

    def goto_home_page(self):
        self.root.destroy()
        present_Homepage_gui_frame()

    @staticmethod
    def refresh():
        restart_program()

    waterbill_data_list = Bill_database_read().load_workbook_values()
    total_rows = len(waterbill_data_list)
    total_columns = len(waterbill_data_list[0])


class Gas_Bill:
    def __init__(self, root):
        self.root = root
        Label(root, text="        ", bg="#CAE9F5", fg='black').grid(row=0, column=0)
        Label(root, text="        ", bg="#CAE9F5", fg='black').grid(row=1, column=0)
        Label(root, text="        ", bg="#CAE9F5", fg='black').grid(row=2, column=0)

        # ------------------Title----------------------
        Label(root, text="Gas Bill", bg="#CAE9F5", fg='Black', font=("Helvetica", 15, 'bold')).place(x=60, y=10)

        refresh_button = Button(root, text="Refresh", bg="#ADECDF", command=self.refresh)
        refresh_button.place(x=360, y=140)

        Back_button = Button(root, text="Back", bg="#ADECDF", command=self.go_back)
        Back_button.place(x=255, y=20)

        Home_button = Button(root, text="Home Page", bg="#ADECDF", command=self.goto_home_page)
        Home_button.place(x=295, y=20)

        quit_button = Button(root, text="Quit", bg="#ADECDF", command=root.destroy)
        quit_button.place(x=360, y=200)

        delete_bill_row_button = Button(root, text="Delete row :", bg="#ADECDF", command=self.delete_bill_row)
        delete_bill_row_button.place(x=400, y=20)
        self.delete_bill_row_value = Entry(width=5)
        self.delete_bill_row_value.place(x=480, y=23)

        for i in range(self.total_rows):
            for j in range(self.total_columns):
                if j == 0 or j == 1 or j == 4:
                    if i == 0:
                        self.e = Label(root, text=str(self.gasbill_data_list[i][j]) + "   ", bg="#CAE9F5", fg='black',
                                       font=("Helvetica", 10, 'bold'))
                        self.e.grid(row=i + 3, column=j)
                    else:
                        if j == 1:
                            self.btn = Button(root, text=self.gasbill_data_list[i][j], width=5)
                            self.btn.grid(row=i + 3, column=j, pady=10)
                        else:
                            self.e = Label(root, text=self.gasbill_data_list[i][j], fg='black').grid(row=i + 3,
                                                                                                     column=j, pady=10)

                if i == self.total_rows - 1:
                    if j == 1:
                        Label(root, text='TOTAL = ', fg='black', bg="#CAE9F5", font=("Helvetica", 9, 'bold'),
                              pady=20).grid(row=i + 5, column=j)
                    elif j == 4:
                        Label(root, text=str(Bill_Calculation().gas_Bill()), bg="#CAE9F5", fg='black',
                              font=("Helvetica", 9, 'bold'),
                              pady=20).grid(row=i + 5, column=4)

    def go_back(self):
        self.root.destroy()
        present_BillSection_gui_frame()

    def delete_bill_row(self):
        row = int(self.delete_bill_row_value.get())
        Bill_database_edit().delete_gas_bill_row(row)
        Bill_database_create().save_excel()
        self.refresh()

    def goto_home_page(self):
        self.root.destroy()
        present_Homepage_gui_frame()

    @staticmethod
    def refresh():
        restart_program()

    gasbill_data_list = Bill_database_read().load_workbook_values()
    total_rows = len(gasbill_data_list)
    total_columns = len(gasbill_data_list[0])


class Add_Bill:
    def __init__(self, master):
        self.master = master

        Label(master, text="                             ", bg="#CAE9F5", font=("Helvetica", 10)).grid(row=0, column=0)
        Label(master, text="Add Bill info-", bg="#CAE9F5", font=("Helvetica", 10, 'bold')).place(x=0, y=0)

        Home_button = Button(master, text="Home Page", bg="#ADECDF", command=self.goto_home_page)
        Home_button.grid(row=10, column=0, sticky=E)

        quit_button = Button(master, text="Quit", bg="#ADECDF", command=master.destroy)
        quit_button.grid(row=10, column=2, sticky=E)

        l1 = Label(master, text="Flat number :", bg="#CAE9F5")
        l2 = Label(master, text="Owner name :", bg="#CAE9F5")
        l3 = Label(master, text="Electric Bill :", bg="#CAE9F5")
        l4 = Label(master, text="Gas Bill:", bg="#CAE9F5")
        l5 = Label(master, text="Water Bill:", bg="#CAE9F5")

        l1.grid(row=1, column=0, sticky=W, pady=2)
        l2.grid(row=2, column=0, sticky=W, pady=2)
        l3.grid(row=3, column=0, sticky=W, pady=2)
        l4.grid(row=4, column=0, sticky=W, pady=2)
        l5.grid(row=5, column=0, sticky=W, pady=2)

        self.e1 = Entry()
        self.e2 = Entry()
        self.e3 = Entry()
        self.e4 = Entry()
        self.e5 = Entry()

        self.e1.grid(row=1, column=1, pady=2)
        self.e2.grid(row=2, column=1, pady=2)
        self.e3.grid(row=3, column=1, pady=2)
        self.e4.grid(row=4, column=1, pady=2)
        self.e5.grid(row=5, column=1, pady=2)

        b1 = Button(master, text="Add", bg="#ADECDF", command=self.add_value)
        b1.grid(row=10, column=1, sticky=E)

    def add_value(self):
        data_list = [self.e1.get(), self.e2.get(), self.e3.get(), self.e4.get(), self.e5.get()]
        Bill_database_edit().add_information(data_list)
        Bill_database_create().save_excel()

        restart_program()

    def goto_home_page(self):
        self.master.destroy()
        present_Homepage_gui_frame()


class Add_Owner:
    def __init__(self, master):
        self.master = master

        Label(master, text="                             ", bg="#CAE9F5", font=("Helvetica", 10)).grid(row=0, column=0)
        Label(master, text="Add flat owner info-", bg="#CAE9F5", font=("Helvetica", 10, 'bold')).place(x=0, y=0)

        Home_button = Button(master, text="Home Page", bg="#ADECDF", command=self.goto_home_page)
        Home_button.grid(row=10, column=0, sticky=E)

        quit_button = Button(master, text="Quit", bg="#ADECDF", command=master.destroy)
        quit_button.grid(row=10, column=2, sticky=E)

        l1 = Label(master, text="Flat number :", bg="#CAE9F5")
        l2 = Label(master, text="Owner name :", bg="#CAE9F5")
        l3 = Label(master, text="Contact Number :", bg="#CAE9F5")
        l4 = Label(master, text="Intercom no. :", bg="#CAE9F5")
        l5 = Label(master, text="Parking :", bg="#CAE9F5")

        l1.grid(row=1, column=0, sticky=W, pady=2)
        l2.grid(row=2, column=0, sticky=W, pady=2)
        l3.grid(row=3, column=0, sticky=W, pady=2)
        l4.grid(row=4, column=0, sticky=W, pady=2)
        l5.grid(row=5, column=0, sticky=W, pady=2)

        self.e1 = Entry()
        self.e2 = Entry()
        self.e3 = Entry()
        self.e4 = Entry()
        self.e5 = Entry()

        self.e1.grid(row=1, column=1, pady=2)
        self.e2.grid(row=2, column=1, pady=2)
        self.e3.grid(row=3, column=1, pady=2)
        self.e4.grid(row=4, column=1, pady=2)
        self.e5.grid(row=5, column=1, pady=2)

        b1 = Button(master, text="Add", bg="#ADECDF", command=self.add_value)
        b1.grid(row=10, column=1, sticky=E)

    def add_value(self):
        data_list = []
        data_list2 = []
        data_list.append(self.e1.get())
        data_list2.append(self.e1.get())
        data_list.append(self.e2.get())
        data_list2.append(self.e2.get())
        data_list.append(self.e3.get())
        data_list.append(self.e4.get())
        data_list.append(self.e5.get())
        Flat_database_edit().add_information(data_list)
        Flat_database_create().save_excel()

        Bill_database_edit().add_information(data_list2)
        Bill_database_create().save_excel()

        restart_program()

    def goto_home_page(self):
        self.master.destroy()
        present_Homepage_gui_frame()


class UserManual:
    def __init__(self, window):
        self.window = window
        Title = Label(window, text="USER MANUAL", fg='Black', bg="#CAE9F5", font=("Helvetica", 40, 'bold'))
        Title.place(x=220, y=10)

        Label(window, text="01. Add titles for both new files 'Flat_Database' & 'Bill_database'. Click the above right buttons to create titles in the database.",
              fg='Black', bg="#CAE9F5", font=("Helvetica", 10)).place(x=10, y=100)

        Label(window, text="02. Click Flat Profile to view the list of flats in the Apartment with detailed information.",
              fg='Black', bg="#CAE9F5", font=("Helvetica", 10)).place(x=10, y=130)

        Label(window, text="03. Click Bill section to view the paid and unpaid Bill database for each flat and total monthly bill for each flat and whole apartment.",
              fg='Black', bg="#CAE9F5", font=("Helvetica", 10)).place(x=10, y=160)

        Label(window, text="04. Click Add Members to include an apartment with the owner's and flat's information.",
              fg='Black', bg="#CAE9F5", font=("Helvetica", 10)).place(x=10, y=190)

        Label(window, text="05. Click Add Bill to update bill data for each flat according to flat number.",
              fg='Black', bg="#CAE9F5", font=("Helvetica", 10)).place(x=10, y=220)

        Label(window, text="06. Click delete row button in flat profile to remove a complete data of a flat according to the selected row number.",
              fg='Black', bg="#CAE9F5", font=("Helvetica", 10)).place(x=10, y=250)

        Label(window, text="07. Click clear row button in Bill Section to clear a bill data of a flat according to the selected row number.",
              fg='Black', bg="#CAE9F5", font=("Helvetica", 10)).place(x=10, y=280)

        quit_button = Button(window, text="Quit", bg="#ADECDF", command=window.destroy)
        quit_button.place(x=600, y=400)

        Homepage_button = Button(window, text="Back to Homepage", bg="#ADECDF", command=self.goto_HomePage)
        Homepage_button.place(x=170, y=400)

    def goto_HomePage(self):
        self.window.destroy()
        present_Homepage_gui_frame()


class Homepage:
    def __init__(self, window):
        self.window = window
        # ------------------Title----------------------
        Title = Label(window, text="HOME PAGE", fg='Black', bg="#CAE9F5", font=("Helvetica", 40, 'bold'))
        Title.place(x=220, y=10)

        note = Label(window,
                     text="Add titles for both new files 'Flat_Database' & 'Bill_database'. Click the above right buttons to create titles in the database",
                     fg='Black', bg="#CAE9F5", font=("Helvetica", 10))
        note.place(x=10, y=400)

        create_Flat_data_button = Button(window, text="Create title for flat profile", bg="#ADECDF",
                                         command=self.createFlatDatabase)
        create_Flat_data_button.place(x=400, y=200)

        create_bill_data_button = Button(window, text="Create title for bill section", bg="#ADECDF",
                                         command=self.createBilldatabase)
        create_bill_data_button.place(x=400, y=240)

        quit_button = Button(window, text="Quit", bg="#ADECDF", command=window.destroy)
        quit_button.place(x=400, y=300)

        # -----------------Sections---------------------
        lbl = Label(window, text="Select a database to visit in the list below:", bg="#CAE9F5", fg='Black',
                    font=("Helvetica", 14, 'bold'))
        lbl.place(x=10, y=100)

        flat_Profile_btn = Button(window, text="Flat Profile", bg="#ADECDF", fg='blue', command=self.goto_flatProfile)
        flat_Profile_btn.place(x=20, y=140)

        Bill_Section_btn = Button(window, text="Bill Section", bg="#ADECDF", fg='blue', command=self.goto_billSection)
        Bill_Section_btn.place(x=20, y=180)

        Add_bill_btn = Button(window, text="Add Bill", bg="#ADECDF", fg='blue', command=self.goto_AddBill)
        Add_bill_btn.place(x=20, y=220)

        Add_flat_members_btn = Button(window, text="Add flat members", bg="#ADECDF", fg='blue',
                                      command=self.goto_AddMember)
        Add_flat_members_btn.place(x=20, y=260)

        Service_Charge_btn = Button(window, text="Service Charge", bg="#ADECDF", fg='blue')
        Service_Charge_btn.place(x=20, y=300)

        UserManual_btn = Button(window, text="User Manual", bg="#D16A9A", fg='white', command=self.goto_userManual)
        UserManual_btn.place(x=20, y=340)

    @staticmethod
    def createFlatDatabase():
        Flat_database_create().add_title()
        Flat_database_create().save_excel()

    @staticmethod
    def createBilldatabase():
        Bill_database_create().add_title()
        Bill_database_create().save_excel()

    def goto_flatProfile(self):
        self.window.destroy()
        present_FlatProfile_gui_frame()

    def goto_billSection(self):
        self.window.destroy()
        present_BillSection_gui_frame()

    def goto_AddMember(self):
        self.window.destroy()
        present_AddOwner_gui_frame()

    def goto_AddBill(self):
        self.window.destroy()
        present_AddBill_gui_frame()

    def goto_userManual(self):
        self.window.destroy()
        present_UserManual_gui_frame()


def present_Homepage_gui_frame():
    window = Tk()
    Homepage(window)
    windowWidth = window.winfo_reqwidth()
    windowHeight = window.winfo_reqheight()

    positionRight = int(window.winfo_screenwidth() / 3 - windowWidth / 2)
    positionDown = int(window.winfo_screenheight() / 3 - windowHeight / 2)

    window.title('Apartment Management System')
    window.geometry("840x450+{}+{}".format(positionRight, positionDown))
    window['background'] = '#CAE9F5'
    window.mainloop()


def present_UserManual_gui_frame():
    window = Tk()
    UserManual(window)
    windowWidth = window.winfo_reqwidth()
    windowHeight = window.winfo_reqheight()

    positionRight = int(window.winfo_screenwidth() / 3 - windowWidth / 2)
    positionDown = int(window.winfo_screenheight() / 3 - windowHeight / 2)

    window.title('Apartment Management System')
    window.geometry("840x450+{}+{}".format(positionRight, positionDown))
    window['background'] = '#CAE9F5'
    window.mainloop()


def present_AddOwner_gui_frame():
    master = Tk()
    Add_Owner(master)

    windowWidth = master.winfo_reqwidth()
    windowHeight = master.winfo_reqheight()

    positionRight = int(master.winfo_screenwidth() / 2 - windowWidth / 2)
    positionDown = int(master.winfo_screenheight() / 2 - windowHeight / 2)

    master.title("Apartment Management System")
    master.geometry("400x200+{}+{}".format(positionRight, positionDown))
    master['background'] = '#CAE9F5'
    mainloop()


def present_AddBill_gui_frame():
    master = Tk()
    Add_Bill(master)

    windowWidth = master.winfo_reqwidth()
    windowHeight = master.winfo_reqheight()

    positionRight = int(master.winfo_screenwidth() / 2 - windowWidth / 2)
    positionDown = int(master.winfo_screenheight() / 2 - windowHeight / 2)

    master.title("Apartment Management System")
    master.geometry("320x200+{}+{}".format(positionRight, positionDown))
    master['background'] = '#CAE9F5'
    mainloop()


def present_BillSection_gui_frame():
    root = Tk()
    Bill_Section(root)

    windowWidth = root.winfo_reqwidth()
    windowHeight = root.winfo_reqheight()

    positionRight = int(root.winfo_screenwidth() / 3 - windowWidth / 2)
    positionDown = int(root.winfo_screenheight() / 3 - windowHeight / 2)

    root.title('Apartment Management System')
    root.geometry("950x500+{}+{}".format(positionRight, positionDown))
    root['background'] = '#CAE9F5'
    root.mainloop()


def present_FlatProfile_gui_frame():
    root = Tk()
    FlatProfile(root)

    windowWidth = root.winfo_reqwidth()
    windowHeight = root.winfo_reqheight()

    positionRight = int(root.winfo_screenwidth() / 3 - windowWidth / 2)
    positionDown = int(root.winfo_screenheight() / 3 - windowHeight / 2)

    root.title('Apartment Management System')
    root.geometry("860x350+{}+{}".format(positionRight, positionDown))
    root['background'] = '#CAE9F5'
    root.mainloop()


def present_waterBill_gui_frame():
    root = Tk()
    Water_bill(root)

    windowWidth = root.winfo_reqwidth()
    windowHeight = root.winfo_reqheight()

    positionRight = int(root.winfo_screenwidth() / 3 - windowWidth / 2)
    positionDown = int(root.winfo_screenheight() / 3 - windowHeight / 2)

    root.title('Apartment Management System')
    root.geometry("550x500+{}+{}".format(positionRight, positionDown))
    root['background'] = '#CAE9F5'
    root.mainloop()


def present_gasBill_gui_frame():
    root = Tk()
    Gas_Bill(root)

    windowWidth = root.winfo_reqwidth()
    windowHeight = root.winfo_reqheight()

    positionRight = int(root.winfo_screenwidth() / 3 - windowWidth / 2)
    positionDown = int(root.winfo_screenheight() / 3 - windowHeight / 2)

    root.title('Apartment Management System')
    root.geometry("550x500+{}+{}".format(positionRight, positionDown))
    root['background'] = '#CAE9F5'
    root.mainloop()


def present_electricBill_gui_frame():
    root = Tk()
    Electric_Bill(root)

    windowWidth = root.winfo_reqwidth()
    windowHeight = root.winfo_reqheight()

    positionRight = int(root.winfo_screenwidth() / 3 - windowWidth / 2)
    positionDown = int(root.winfo_screenheight() / 3 - windowHeight / 2)

    root.title('Apartment Management System')
    root.geometry("550x500+{}+{}".format(positionRight, positionDown))
    root['background'] = '#CAE9F5'
    root.mainloop()


# present_Homepage_gui_frame()
